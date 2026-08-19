"""Views for Reports and Data Export (CSV / Excel)."""

import csv
import io
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Sum, Q
from rest_framework.views import APIView
from rest_framework import permissions, status

from apps.expenses.models import Expense
from apps.income.models import Income
from core.exceptions import StandardResponse

# Optional Excel library openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def format_export_date(date_val):
    """Format date to clean readable 'DD Mon YYYY' (e.g. 19 Aug 2026) for CSV & Excel."""
    if not date_val:
        return ""
    if hasattr(date_val, 'strftime'):
        return date_val.strftime('%d %b %Y')
    if isinstance(date_val, str) and '-' in date_val:
        try:
            dt = datetime.strptime(date_val[:10], '%Y-%m-%d')
            return dt.strftime('%d %b %Y')
        except Exception:
            return str(date_val)
    return str(date_val)


def get_filtered_transactions(user, params):
    """Helper to retrieve and filter combined expenses and incomes using one canonical source of truth."""
    month = params.get('month')
    year = params.get('year')
    start_date = params.get('start_date')
    end_date = params.get('end_date')
    category_id = params.get('category')
    raw_type = params.get('type', '')

    if not raw_type or str(raw_type).strip().upper() in ['ALL', '']:
        transaction_type = 'ALL'
    else:
        transaction_type = str(raw_type).strip().upper()

    expenses_qs = Expense.objects.filter(user=user, is_deleted=False).select_related('category')
    incomes_qs = Income.objects.filter(user=user, is_deleted=False).select_related('category')

    # Apply date filters
    if start_date and str(start_date).strip():
        expenses_qs = expenses_qs.filter(date__gte=start_date)
        incomes_qs = incomes_qs.filter(date__gte=start_date)
    if end_date and str(end_date).strip():
        expenses_qs = expenses_qs.filter(date__lte=end_date)
        incomes_qs = incomes_qs.filter(date__lte=end_date)
    if month and str(month).strip():
        try:
            m_int = int(month)
            expenses_qs = expenses_qs.filter(date__month=m_int)
            incomes_qs = incomes_qs.filter(date__month=m_int)
        except (ValueError, TypeError):
            pass
    if year and str(year).strip():
        try:
            y_int = int(year)
            expenses_qs = expenses_qs.filter(date__year=y_int)
            incomes_qs = incomes_qs.filter(date__year=y_int)
        except (ValueError, TypeError):
            pass
    if category_id and str(category_id).strip():
        expenses_qs = expenses_qs.filter(category_id=category_id)
        incomes_qs = incomes_qs.filter(category_id=category_id)

    records = []
    if transaction_type in ['ALL', 'EXPENSE']:
        for exp in expenses_qs:
            records.append({
                'id': f"exp-{exp.id}",
                'raw_id': exp.id,
                'type': 'EXPENSE',
                'category': exp.category.name,
                'category_id': exp.category.id,
                'amount': exp.amount,
                'date': exp.date,
                'payment_method': exp.payment_method,
                'description': exp.description or '',
                'notes': exp.notes or '',
                'created_at': exp.created_at
            })

    if transaction_type in ['ALL', 'INCOME']:
        for inc in incomes_qs:
            records.append({
                'id': f"inc-{inc.id}",
                'raw_id': inc.id,
                'type': 'INCOME',
                'category': inc.category.name,
                'category_id': inc.category.id,
                'amount': inc.amount,
                'date': inc.date,
                'payment_method': inc.payment_method,
                'description': inc.description or '',
                'notes': inc.notes or '',
                'created_at': inc.created_at
            })

    # Sort descending by transaction date then created_at
    records.sort(key=lambda x: (x['date'], x['created_at']), reverse=True)
    return records


class ReportSummaryView(APIView):
    """
    Returns filtered report summaries and transactions table.
    
    GET /api/reports/summary/
    """
    
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        records = get_filtered_transactions(request.user, request.query_params)
        
        total_income = sum((r['amount'] for r in records if r['type'] == 'INCOME'), Decimal('0.00'))
        total_expense = sum((r['amount'] for r in records if r['type'] == 'EXPENSE'), Decimal('0.00'))
        net_balance = total_income - total_expense

        serialized_records = []
        for r in records:
            formatted_date = r['date'].strftime('%Y-%m-%d') if hasattr(r['date'], 'strftime') else str(r['date'])
            serialized_records.append({
                'id': r['id'],
                'type': r['type'],
                'category': r['category'],
                'category_id': r['category_id'],
                'amount': str(r['amount']),
                'date': formatted_date,
                'payment_method': r['payment_method'],
                'description': r['description'],
                'notes': r['notes']
            })

        return StandardResponse.success(
            data={
                'total_income': str(total_income),
                'total_expense': str(total_expense),
                'net_balance': str(net_balance),
                'net_savings': str(net_balance),
                'count': len(records),
                'transaction_count': len(records),
                'transactions': serialized_records
            },
            message="Report summary generated successfully."
        )


class ExportCSVView(APIView):
    """
    Generates and streams a CSV export of filtered transactions with actual transaction dates.
    
    GET /api/reports/export/csv/
    """
    
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        records = get_filtered_transactions(request.user, request.query_params)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"expense_tracker_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Type', 'Category', 'Description', 'Payment Method', 'Amount (INR)'])

        for r in records:
            formatted_date = format_export_date(r['date'])
            formatted_amount = f"-{r['amount']}" if r['type'] == 'EXPENSE' else str(r['amount'])
            writer.writerow([
                formatted_date,
                r['type'],
                r['category'],
                r['description'],
                r['payment_method'],
                formatted_amount
            ])

        return response


class ExportExcelView(APIView):
    """
    Generates and returns an Excel (.xlsx) workbook of filtered transactions.
    
    GET /api/reports/export/excel/
    """
    
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        records = get_filtered_transactions(request.user, request.query_params)
        filename = f"expense_tracker_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV format if openpyxl not installed
            return ExportCSVView().get(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions Report"
        ws.views.sheetView[0].showGridLines = True

        # Header Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border_side = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        headers = ['Date', 'Type', 'Category', 'Description', 'Payment Method', 'Amount (INR)']
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        income_font = Font(color="10B981", bold=True)
        expense_font = Font(color="EF4444", bold=True)

        # Write Data
        for row_idx, r in enumerate(records, start=2):
            formatted_date = format_export_date(r['date'])
            signed_amount = float(-r['amount'] if r['type'] == 'EXPENSE' else r['amount'])

            ws.append([
                formatted_date,
                r['type'],
                r['category'],
                r['description'],
                r['payment_method'],
                signed_amount
            ])
            
            # Format type column color
            type_cell = ws.cell(row=row_idx, column=2)
            type_cell.font = income_font if r['type'] == 'INCOME' else expense_font
            type_cell.alignment = Alignment(horizontal="center")

            # Format amount column
            amount_cell = ws.cell(row=row_idx, column=6)
            amount_cell.number_format = '₹#,##0.00;[Red]-₹#,##0.00;₹0.00'
            amount_cell.alignment = Alignment(horizontal="right")

            # Border for all cells
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = cell_border

        # Adjust Column Widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
